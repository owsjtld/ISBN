import os
import re
import time
import requests
import pandas as pd
import openpyxl
from io import BytesIO
from dotenv import load_dotenv

load_dotenv()

DEFAULT_API_KEY = os.getenv("KAKAO_API_KEY", "")


def _clean_title(title: str) -> str:
    return re.sub(r'\(.*?\)|\[.*?\]', '', str(title)).strip()


def _clean_isbn(x) -> str:
    s = str(x).strip()
    if not s or s in ('nan', 'None', ''):
        return ''
    try:
        return str(int(float(s)))
    except Exception:
        return s


SEARCH_MODES = {
    "제목 + 출판사 → 실패 시 제목만 재시도": "both_fallback",
    "제목 + 출판사만": "title_pub",
    "제목만": "title_only",
}


def fetch_isbn(title: str, publisher: str, api_key: str, search_mode: str = "both_fallback") -> str:
    cleaned = _clean_title(title)
    url = "https://dapi.kakao.com/v3/search/book"
    headers = {"Authorization": f"KakaoAK {api_key}"}

    if search_mode == "title_only":
        queries = [cleaned]
    elif search_mode == "title_pub":
        queries = [f"{cleaned} {publisher}"]
    else:  # both_fallback
        queries = [f"{cleaned} {publisher}", cleaned]

    for query in queries:
        try:
            res = requests.get(url, headers=headers, params={"query": query}, timeout=5)
            if res.status_code == 200:
                docs = res.json().get('documents', [])
                if docs:
                    return str(docs[0]['isbn'].split()[-1])
        except Exception:
            pass
    return ""


def process_dataframe(
    df: pd.DataFrame,
    api_key: str,
    search_mode: str = "both_fallback",
    progress_callback=None,
) -> pd.DataFrame:
    """
    ISBN을 채워 넣은 새 DataFrame을 반환한다.
    progress_callback(current, total, title, isbn, success, skipped)
    """
    df = df.copy()
    if 'ISBN' not in df.columns:
        df['ISBN'] = ""
    df['ISBN'] = df['ISBN'].fillna("").astype(str)

    total = len(df)

    for pos, (index, row) in enumerate(df.iterrows(), start=1):
        current_isbn = str(row['ISBN']).strip()

        if len(current_isbn) >= 10:
            if progress_callback:
                progress_callback(pos, total, str(row['도서명']), current_isbn, True, True)
            continue

        title = str(row['도서명'])
        pub = str(row.get('출판사', ''))
        isbn = fetch_isbn(title, pub, api_key, search_mode)

        if isbn:
            df.at[index, 'ISBN'] = isbn

        if progress_callback:
            progress_callback(pos, total, title, isbn, bool(isbn), False)

        time.sleep(0.1)

    df['ISBN'] = df['ISBN'].apply(_clean_isbn)
    return df


def to_excel_bytes(df: pd.DataFrame) -> bytes:
    """ISBN 컬럼을 텍스트 서식으로 저장한 엑셀 바이트 반환"""
    buf = BytesIO()
    df.to_excel(buf, index=False)
    buf.seek(0)

    wb = openpyxl.load_workbook(buf)
    ws = wb.active
    isbn_col = next((c.column for c in ws[1] if c.value == 'ISBN'), None)
    if isbn_col:
        for (cell,) in ws.iter_rows(min_row=2, min_col=isbn_col, max_col=isbn_col):
            if cell.value is not None:
                cell.value = str(cell.value)
                cell.number_format = '@'

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
